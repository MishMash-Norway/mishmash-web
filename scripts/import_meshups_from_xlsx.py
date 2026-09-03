#!/usr/bin/env python3
"""Import MeshUp events from an XLSX schedule.

Preview changes by default. Pass --write to update event files. Portrait mappings
copy a source image into site/assets/images/portraits and update the event image:

    python3 scripts/import_meshups_from_xlsx.py --write \
      --portrait 23=temp/Vinoo_Alluri.jpeg
"""
from __future__ import annotations

import argparse
from datetime import date, datetime
import json
from pathlib import Path
import re
import shutil
import sys

try:
    from openpyxl import load_workbook
except ImportError as error:
    raise SystemExit("openpyxl is required for .xlsx imports (pip install openpyxl)") from error


REQUIRED_HEADERS = {"Date", "WP", "Moderator", "Speaker", "URL to speaker", "Title", "Abstract", "Bio"}
ACCESS_TEXT = """## Access

MishMash MeetUps are short, informal meetings in the consortium where both early career and established researchers present ongoing projects. The events are open for everyone, but, for security reasons, Zoom links are only provided to people that are affiliated with a MishMash Work Package. If not, please [ask for access](mailto:contact@mishmash.no).
"""
DEFAULT_IMAGE = "/assets/images/bubbles/mishmash_bubbles_notext.svg"


def parse_args(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=Path("temp/MishMash Meetups.xlsx"), help="Source workbook")
    parser.add_argument("--write", action="store_true", help="Write event files and copy portraits")
    parser.add_argument("--include-past", action="store_true", help="Also import events dated before today")
    parser.add_argument(
        "--portrait",
        action="append",
        default=[],
        metavar="NUMBER=PATH",
        help="Copy PATH into portraits and use it for MeshUp NUMBER (repeatable)",
    )
    return parser.parse_args(argv)


def clean(value):
    return "" if value is None else str(value).strip()


def parse_portraits(values):
    portraits = {}
    for value in values:
        number, separator, source = value.partition("=")
        if not separator or not number.isdigit() or not source:
            raise ValueError(f"Invalid --portrait {value!r}; use NUMBER=PATH")
        portraits[int(number)] = Path(source)
    return portraits


def parse_event_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.fromisoformat(clean(value)).date()
    except ValueError:
        return None


def read_events(xlsx):
    workbook = load_workbook(xlsx, data_only=True, read_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    missing_headers = REQUIRED_HEADERS - set(headers)
    if missing_headers:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing_headers))}")
    columns = {header: index for index, header in enumerate(headers)}
    events = []
    for row in rows:
        row = row + (None,) * (len(headers) - len(row))
        number_value = row[0] if row else None
        if not isinstance(number_value, (int, float)):
            continue
        if clean(row[columns["Moderator"]]).lower().startswith("skipping"):
            continue
        event_date = parse_event_date(row[columns["Date"]])
        if event_date is None:
            continue
        number = int(number_value)
        events.append({header: clean(row[index]) for header, index in columns.items()} | {"number": number, "date": event_date})
    numbers = [event["number"] for event in events]
    duplicates = sorted(number for number in set(numbers) if numbers.count(number) > 1)
    if duplicates:
        raise ValueError(f"Duplicate MeshUp number(s): {', '.join(map(str, duplicates))}")
    return events


def markdown_link(name, url):
    return f"[{name}]({url})" if url.startswith(("https://", "http://")) else name


def event_content(event, image):
    speaker = event["Speaker"]
    title = event["Title"]
    topic_is_known = bool(title and title.upper() != "TBD")
    heading = f"MeshUp #{event['number']} - {title}" if topic_is_known else f"MeshUp #{event['number']} - Weekly MishMash Gathering"
    description = (
        f"{speaker} presents {title}." if speaker and topic_is_known
        else f"{speaker} will give MeshUp #{event['number']}. Title and abstract coming soon." if speaker
        else f"Placeholder for MeshUp #{event['number']}. Speaker and topic to be announced."
    )
    lines = [
        "---",
        f"title: {json.dumps(heading, ensure_ascii=False)}",
        f"date: {event['date'].isoformat()} 12:00:00 +02:00",
        f"end_date: {event['date'].isoformat()} 12:30:00 +02:00",
        "location: Zoom",
        "layout: event",
        f"categories: [MeshUp, WP{event['WP']} ]".replace(" ]", "]"),
        "tags: []",
        f"description: {json.dumps(description, ensure_ascii=False)}",
        f"image: {image}",
        f"slug: \"meshup-{event['number']}\"",
        "---",
        "",
    ]
    if speaker:
        presentation = markdown_link(speaker, event["URL to speaker"])
        lines.append(f"This week, {presentation} will present{' *' + title + '*' if topic_is_known else ''}.")
    else:
        lines.append(f"Details for MeshUp #{event['number']} will be announced soon.")
    if event["Abstract"] and event["Abstract"].upper() != "TBD":
        lines.extend(["", "## Abstract", event["Abstract"]])
    if event["Bio"]:
        lines.extend(["", "## Bio", event["Bio"]])
    lines.extend(["", ACCESS_TEXT])
    return "\n".join(lines)


def existing_image(path):
    if not path.exists():
        return DEFAULT_IMAGE
    match = re.search(r"^image:\s*(\S+)", path.read_text(encoding="utf-8"), re.MULTILINE)
    return match.group(1) if match else DEFAULT_IMAGE


def main(argv=None):
    args = parse_args(argv)
    if not args.xlsx.exists():
        print(f"XLSX not found: {args.xlsx}", file=sys.stderr)
        return 2
    try:
        portraits = parse_portraits(args.portrait)
        events = read_events(args.xlsx)
    except ValueError as error:
        print(error, file=sys.stderr)
        return 2
    if not args.include_past:
        events = [event for event in events if event["date"] >= date.today()]

    repo_root = Path(__file__).resolve().parents[1]
    event_directory = repo_root / "site" / "_events"
    portrait_directory = repo_root / "site" / "assets" / "images" / "portraits"
    existing_paths = {
        int(match.group(1)): path
        for path in event_directory.glob("*-meshup*.md")
        if (match := re.search(r"-meshup(\d+)\.md$", path.name))
    }
    changes = 0
    for event in events:
        number = event["number"]
        event_path = event_directory / f"{event['date'].isoformat()}-meshup{number:02d}.md"
        existing_path = existing_paths.get(number, event_path)
        image = existing_image(existing_path)
        source = portraits.get(number)
        if source:
            if not source.exists():
                print(f"Portrait not found: {source}", file=sys.stderr)
                return 2
            image = f"/assets/images/portraits/{source.name}"
        content = event_content(event, image)
        current = existing_path.read_text(encoding="utf-8") if existing_path.exists() else ""
        if content == current and existing_path == event_path:
            continue
        changes += 1
        action = "Update" if existing_path.exists() else "Create"
        print(f"{action} {event_path.relative_to(repo_root)}")
        if args.write:
            event_path.write_text(content, encoding="utf-8")
            if existing_path != event_path and existing_path.exists():
                existing_path.unlink()
            if source:
                portrait_directory.mkdir(parents=True, exist_ok=True)
                shutil.copy2(source, portrait_directory / source.name)
    print(f"{changes} event file(s) {'written' if args.write else 'would be written'}.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())