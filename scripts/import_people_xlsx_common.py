#!/usr/bin/env python3
"""Shared helpers for XLSX-based people import scripts."""

from __future__ import annotations

from copy import deepcopy
from pathlib import Path
import re
import sys
import unicodedata

try:
    from openpyxl import load_workbook
except Exception:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    raise

from directory_io import load_entry, save_entry


INCLUDE_CANDIDATES = [
    "include",
    "include in directory",
    "do you want to be added to the mishmash directory",
    "added to the mishmash directory",
    "add to directory",
    "include?",
    "in directory",
    "include_flag",
    "add_flag",
    "publish",
    "published",
    "add_to_directory",
]

URL_FIELD_ALIASES = {
    "personal_website": ["web page (personal)", "website (personal)", "personal website"],
    "institutional_website": ["web page (institutional)", "website (institution)", "institutional website"],
    "github": ["github"],
    "linkedin": ["linkedin"],
    "youtube": ["youtube"],
    "facebook": ["facebook"],
    "mastodon": ["mastodon"],
    "instagram": ["instagram"],
    "bluesky": ["bluesky"],
    "orcid": ["orcid"],
    "nva": ["nva"],
}


def canonical_orcid_url(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return ""
    # Accept any input containing an ORCID iD (bare, path, or e.g. a
    # my-orcid?orcid=... dashboard URL) and canonicalize it.
    match = re.search(r"(\d{4}-\d{4}-\d{4}-[\dX]{4})", value, flags=re.IGNORECASE)
    if match:
        return f"https://orcid.org/{match.group(1).upper()}"
    return value


def canonical_nva_url(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return ""
    match = re.search(r"research-profile/(\d+)", value)
    if match:
        return f"https://nva.sikt.no/research-profile/{match.group(1)}"
    if value.isdigit():
        return f"https://nva.sikt.no/research-profile/{value}"
    if value.startswith("https://nva.sikt.no/") or value.startswith("http://nva.sikt.no/"):
        # nva.sikt.no URLs without a research-profile id (my-page dashboards,
        # search links) identify nothing; drop them so they cannot clobber a
        # previously stored profile URL.
        return ""
    return ""


def normalize_http_url(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return ""
    if value.startswith("http://"):
        return "https://" + value.removeprefix("http://")
    if value.startswith("https://"):
        return value
    if re.match(r"^[a-zA-Z][a-zA-Z0-9+.-]*:", value):
        return value
    return f"https://{value.lstrip('/')}"


def first_url(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return ""
    for part in re.split(r"[\s,|]+", value):
        part = part.strip()
        if not part:
            continue
        if "." in part or part.startswith("http"):
            return normalize_http_url(part)
    return normalize_http_url(value)


def normalize_field_value(field_name: str, value: str) -> str:
    value = (value or "").strip()
    if not value:
        return ""
    if field_name == "orcid":
        return canonical_orcid_url(value)
    if field_name == "nva":
        return canonical_nva_url(value)
    if field_name in {"personal_website", "institutional_website", "github", "linkedin", "youtube", "facebook"}:
        return first_url(value)
    if field_name == "instagram":
        cleaned = value.removeprefix("@").strip().strip("/")
        if "instagram.com" in cleaned:
            return first_url(cleaned)
        return f"https://www.instagram.com/{cleaned}/" if cleaned else ""
    if field_name == "mastodon":
        cleaned = value.strip().removeprefix("@").strip().strip("/")
        if "@" in cleaned and not cleaned.startswith("http"):
            handle, instance = cleaned.split("@", 1)
            if handle and instance:
                return f"https://{instance}/@{handle}"
        return first_url(cleaned)
    if field_name == "bluesky":
        cleaned = value.removeprefix("@").strip().strip("/")
        if "bsky.app/profile/" in cleaned:
            return first_url(cleaned)
        if cleaned.endswith(".bsky.social") or ".bsky.social/" in cleaned:
            handle = cleaned.split("/", 1)[0]
            return f"https://bsky.app/profile/{handle}"
        return first_url(cleaned)
    return value


def canonical_website_url(value: str) -> str:
    return normalize_http_url(value).rstrip("/") if value else ""


def dedupe_website_pair(urls: dict) -> bool:
    personal = canonical_website_url(urls.get("personal_website") or "")
    institutional = canonical_website_url(urls.get("institutional_website") or "")
    if personal and institutional and personal == institutional:
        if urls.get("institutional_website") != "":
            urls["institutional_website"] = ""
            return True
    return False


def slugify(value: str) -> str:
    # NFKD leaves ø/æ (and their upper-case forms) undecomposed, so fold them
    # explicitly before stripping to ASCII; otherwise Løve becomes lve.
    value = value.translate(str.maketrans({"ø": "o", "Ø": "O", "æ": "ae", "Æ": "Ae", "ß": "ss"}))
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    value = re.sub(r"[^\w\s-]", "", value).strip().lower()
    value = re.sub(r"[-\s]+", "-", value)
    return value[:80].strip("-")


def fix_name_case(name: str) -> str:
    """Title-case names submitted in all caps or all lower case.

    Mixed-case names are left untouched so curated forms like
    "von Arnim" or "de Seta" are never mangled.
    """
    if name and (name == name.upper() or name == name.lower()):
        return " ".join(part.capitalize() for part in name.split())
    return name


def truthy(value) -> bool:
    if value is None:
        return False
    return str(value).strip().lower() in {"y", "yes", "true", "1", "x", "include"}


def find_column(headers, candidates):
    for candidate in candidates:
        for index, header in enumerate(headers):
            if header and candidate in header:
                return index
    return None


def normalize_header(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def cell_text(row, index):
    if index is None or index >= len(row):
        return ""
    value = row[index]
    if value is None:
        return ""
    return str(value).strip()


def find_first_value(row, headers, aliases):
    index = find_column(headers, aliases)
    return cell_text(row, index)


def read_people(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return [], "new"

    headers = [normalize_header(cell) for cell in rows[0]]
    name_idx = find_column(headers, ["name", "full name", "display name"]) or 0
    include_idx = find_column(headers, INCLUDE_CANDIDATES)
    sheet_kind = "new" if include_idx is not None else "existing"

    people = []
    for row in rows[1:]:
        if not row or all(cell is None for cell in row):
            continue
        name = fix_name_case(cell_text(row, name_idx))
        if not name:
            continue

        if sheet_kind == "new" and not truthy(row[include_idx]):
            continue

        urls = {}
        for field_name, aliases in URL_FIELD_ALIASES.items():
            value = find_first_value(row, headers, aliases)
            if value:
                normalized = normalize_field_value(field_name, value)
                if normalized:
                    urls[field_name] = normalized

        people.append(
            {
                "name": name,
                "slug": slugify(name),
                "urls": urls,
                "sheet_kind": sheet_kind,
            }
        )

    for person in people:
        dedupe_website_pair(person["urls"])

    return people, sheet_kind


# Fields that identify a person; a stored value is never silently replaced
# by a different one from a form row (cumulative exports resubmit old
# mistakes on every import).
IDENTITY_URL_FIELDS = ("orcid", "nva")


def apply_person_to_entry(data: dict, person: dict, is_new: bool = True, warnings: list | None = None) -> dict:
    updated = deepcopy(data)
    updated["slug"] = person["slug"]
    if is_new or not str(updated.get("name") or "").strip():
        updated["name"] = person["name"]
        updated["title"] = person["name"]
    urls = updated.setdefault("urls", {})
    for field_name, value in person.get("urls", {}).items():
        if not value:
            continue
        normalized = normalize_field_value(field_name, value)
        if not normalized:
            continue
        existing = str(urls.get(field_name) or "").strip()
        if (
            not is_new
            and field_name in IDENTITY_URL_FIELDS
            and existing
            and existing != normalized
        ):
            if warnings is not None:
                warnings.append(
                    f"{person['slug']}: kept {field_name} {existing} (form submitted {normalized})"
                )
            continue
        urls[field_name] = normalized
    for field_name, value in list(urls.items()):
        normalized = normalize_field_value(field_name, value)
        if normalized:
            urls[field_name] = normalized
    dedupe_website_pair(urls)
    return updated


def build_alias_map(out_base: Path) -> dict:
    """Map slugified names and aliases of existing entries to their slug.

    Lets a form row that uses a nickname or short name (e.g. "Shayan" for
    shayan-dadman) update the existing entry instead of creating a duplicate.
    """
    alias_map = {}
    if not out_base.is_dir():
        return alias_map
    for child in sorted(out_base.iterdir()):
        index_md = child / "index.md"
        if child.name.startswith("_") or not index_md.exists():
            continue
        try:
            data, _ = load_entry(index_md)
        except ValueError:
            continue
        candidates = [str(data.get("name") or "")] + [str(a) for a in data.get("aliases") or []]
        for candidate in candidates:
            key = slugify(candidate)
            if key and key != child.name:
                alias_map[key] = child.name
    return alias_map


def drop_duplicate_identity_values(people, warnings: list) -> None:
    """Remove orcid/nva values submitted by rows for different people.

    Two people cannot share an ORCID or NVA profile, so such a value is a
    copy-paste mistake in at least one row; it is applied to neither.
    Multiple rows from the same person (common in cumulative exports) are
    fine and left alone.
    """
    for field_name in IDENTITY_URL_FIELDS:
        owners = {}
        for person in people:
            value = person.get("urls", {}).get(field_name)
            if value:
                owners.setdefault(value, set()).add(person["slug"])
        for person in people:
            value = person.get("urls", {}).get(field_name)
            if value and len(owners[value]) > 1:
                others = sorted(owners[value] - {person["slug"]})
                warnings.append(
                    f"{person['slug']}: dropped {field_name} {value} (also submitted by {', '.join(others)})"
                )
                del person["urls"][field_name]


def import_people(people, template_path: Path, out_base: Path):
    template_data, template_body = load_entry(template_path)
    created = 0
    updated = 0
    warnings = []
    out_base.mkdir(parents=True, exist_ok=True)
    alias_map = build_alias_map(out_base)

    drop_duplicate_identity_values(people, warnings)

    for person in people:
        slug = person["slug"]
        if not slug:
            print(f"Skipping name with empty slug: {person['name']}")
            continue

        if not (out_base / slug / "index.md").exists() and slug in alias_map:
            warnings.append(f"{slug}: matched existing entry {alias_map[slug]} via alias")
            slug = alias_map[slug]
            person = {**person, "slug": slug}

        out_dir = out_base / slug
        out_file = out_dir / "index.md"
        if out_file.exists():
            data, body = load_entry(out_file)
            updated_data = apply_person_to_entry(data, person, is_new=False, warnings=warnings)
            save_entry(out_file, updated_data, body)
            updated += 1
            continue

        out_dir.mkdir(parents=True, exist_ok=True)
        created_data = apply_person_to_entry(template_data, person)
        created_data["permalink"] = f"/people/{slug}/"
        save_entry(out_file, created_data, template_body)
        created += 1

    for warning in warnings:
        print(f"  WARNING: {warning}")

    return created, updated, 0